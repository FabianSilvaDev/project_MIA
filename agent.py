import os
import sys
import re
import tempfile
import subprocess
import io
import wave
import threading
import time
import base64
import platform
import requests
import numpy as np
import asyncio
import websockets
import json
from dotenv import load_dotenv
from rich.console import Console
from rich.live import Live
from rich.panel import Panel
from rich.align import Align
from rich.text import Text


# =========================
# Cargar entorno
# =========================
load_dotenv()
console = Console()

# =========================
# Audio opcional avanzado
# =========================
try:
    import sounddevice  # type: ignore
    _have_sounddevice = True
except Exception:
    _have_sounddevice = False

try:
    import torch  # type: ignore
    from chatterbox.tts import ChatterboxTTS  # type: ignore
    _have_chatterbox = True
except Exception:
    torch = None  # type: ignore
    ChatterboxTTS = None  # type: ignore
    _have_chatterbox = False

# =========================
# WebSocket Server
# =========================

clients = set()
WS_CLIENTS_COUNT = 0
WS_LOOP = None
WS_SERVER_STARTED = False
WS_SERVER_LOCK = threading.Lock()
WS_STATUS_EVENT = threading.Event()
WS_SERVER_HOST = "127.0.0.1"
WS_SERVER_PORT = 3312
CURRENT_EMOTION = "happy"  # Track current emotion for the avatar
WS_DEBUG_LOGS = False
LIPSYNC_OFFSET_MS = int(os.getenv("BMO_LIPSYNC_OFFSET_MS", "0"))

async def register(websocket):
    global WS_CLIENTS_COUNT
    clients.add(websocket)
    WS_CLIENTS_COUNT = len(clients)
    if WS_DEBUG_LOGS:
        print(f"[BMO] Cliente WebSocket conectado (clientes: {WS_CLIENTS_COUNT})")

async def unregister(websocket):
    global WS_CLIENTS_COUNT
    clients.discard(websocket)
    WS_CLIENTS_COUNT = len(clients)
    if WS_DEBUG_LOGS:
        print(f"[BMO] Cliente WebSocket desconectado (clientes: {WS_CLIENTS_COUNT})")

async def broadcast(data):
    if clients:
        await asyncio.gather(
            *[client.send(json.dumps(data)) for client in clients],
            return_exceptions=True
        )

async def handler(websocket):
    await register(websocket)
    try:
        async for message in websocket:
            pass
    finally:
        await unregister(websocket)


def start_websocket_server():
    global WS_LOOP, WS_SERVER_STARTED

    def _is_port_in_use_error(error: OSError) -> bool:
        return (
            getattr(error, "winerror", None) == 10048
            or getattr(error, "errno", None) == 10048
            or "10048" in str(error)
            or "address already in use" in str(error).lower()
        )

    async def _run_server():
        async with websockets.serve(handler, WS_SERVER_HOST, WS_SERVER_PORT):
            WS_STATUS_EVENT.set()
            await asyncio.Future()

    with WS_SERVER_LOCK:
        if WS_SERVER_STARTED:
            WS_STATUS_EVENT.set()
            return
        WS_SERVER_STARTED = True

    loop = asyncio.new_event_loop()
    WS_LOOP = loop
    asyncio.set_event_loop(loop)
    try:
        loop.run_until_complete(_run_server())
    except OSError as e:
        if _is_port_in_use_error(e):
            with WS_SERVER_LOCK:
                WS_SERVER_STARTED = False
            WS_LOOP = None
            WS_STATUS_EVENT.set()
            return
        with WS_SERVER_LOCK:
            WS_SERVER_STARTED = False
        WS_LOOP = None
        WS_STATUS_EVENT.set()
        return
    except Exception:
        with WS_SERVER_LOCK:
            WS_SERVER_STARTED = False
        WS_LOOP = None
        WS_STATUS_EVENT.set()
        return


def send_ws_state(data):
    if WS_LOOP is None:
        return

    try:
        asyncio.run_coroutine_threadsafe(broadcast(data), WS_LOOP)
    except Exception:
        pass


def wait_for_ws_client(timeout_seconds: float = 2.0, poll_interval: float = 0.05) -> bool:
    deadline = time.time() + max(0.0, timeout_seconds)
    while time.time() < deadline:
        if WS_CLIENTS_COUNT > 0:
            return True
        time.sleep(max(0.01, poll_interval))
    return WS_CLIENTS_COUNT > 0


def get_websocket_status() -> dict:
    return {
        "ready": WS_LOOP is not None,
        "started": WS_SERVER_STARTED,
        "host": WS_SERVER_HOST,
        "port": WS_SERVER_PORT,
        "clients": WS_CLIENTS_COUNT,
    }


def detect_emotion_from_text(text: str) -> str:
    """Detecta emoción basada en palabras clave en el texto."""
    text_lower = text.lower()
    
    # Palabras clave para cada emoción
    happy_keywords = ["jaja", "ja ja", "jeje", "bueno", "excelente", "genial", "fantástico", "gracias", ":)", "😊"]
    curious_keywords = ["interesante", "quiero saber", "cuéntame", "cómo", "qué", "interesa", "pregunto", "explora"]
    surprised_keywords = ["wow", "vaya", "sorprendente", "increíble", "nunca", "!"]
    thinking_keywords = ["creo", "parece", "cuestión", "depende", "tal vez", "quizás", "consider"]
    sad_keywords = ["lamento", "lo siento", "triste", "malo", "difícil", "problema", "error", "fallo"]
    
    # Contar coincidencias
    happy_count = sum(1 for kw in happy_keywords if kw in text_lower)
    curious_count = sum(1 for kw in curious_keywords if kw in text_lower)
    surprised_count = sum(1 for kw in surprised_keywords if kw in text_lower)
    thinking_count = sum(1 for kw in thinking_keywords if kw in text_lower)
    sad_count = sum(1 for kw in sad_keywords if kw in text_lower)
    
    # Retornar emoción con más coincidencias
    scores = {
        "happy": happy_count,
        "curious": curious_count,
        "surprised": surprised_count,
        "thinking": thinking_count,
        "sad": sad_count
    }
    
    max_emotion = max(scores, key=scores.get)
    if scores[max_emotion] > 0:
        return max_emotion
    return "neutral"


def _pcm16_bytes_to_float32(audio_bytes: bytes, channels: int = 1):
    pcm = np.frombuffer(audio_bytes, dtype=np.int16)
    if pcm.size == 0:
        return np.empty((0, channels), dtype=np.float32)

    if channels > 1:
        usable = (pcm.size // channels) * channels
        pcm = pcm[:usable].reshape(-1, channels)
    else:
        pcm = pcm.reshape(-1, 1)

    return (pcm.astype(np.float32) / 32768.0)


def stream_audio_to_ws(audio_bytes: bytes, sample_rate: int = 24000, channels: int = 1, start_delay_ms: int = 0):
    if WS_LOOP is None:
        return

    try:
        data = _pcm16_bytes_to_float32(audio_bytes, channels=channels)
        if data.size == 0:
            send_ws_state({"type": "audio_end"})
            return

        chunk_size = 1024
        started_at = time.perf_counter()

        if start_delay_ms > 0:
            time.sleep(start_delay_ms / 1000.0)

        # Asegurar que el frontend sepa que está hablando
        send_ws_state({
            "type": "audio_start",
            "sample_rate": int(sample_rate),
            "channels": int(channels),
        })

        for start in range(0, len(data), chunk_size):
            chunk = data[start:start + chunk_size]
            payload = base64.b64encode(chunk.astype(np.float32).tobytes()).decode("ascii")

            send_ws_state({
                "type": "audio_chunk",
                "format": "pcm_f32le",
                "sample_rate": int(sample_rate),
                "channels": int(channels),
                "data": payload,
            })

            played_seconds = (start + len(chunk)) / sample_rate
            elapsed_seconds = time.perf_counter() - started_at
            sleep_time = played_seconds - elapsed_seconds
            if sleep_time > 0:
                # Mantener tiempo real del stream para que la boca siga el audio.
                time.sleep(sleep_time)

        send_ws_state({"type": "audio_end"})
    except Exception:
        pass


def play_audio_locally(audio_bytes: bytes, sample_rate: int = 24000, channels: int = 1):
    if not _have_sounddevice:
        print("[BMO] Audio local desactivado: falta sounddevice.")
        return

    try:
        data = _pcm16_bytes_to_float32(audio_bytes, channels=channels)
        if data.size == 0:
            return
        sounddevice.play(data, samplerate=sample_rate)
        sounddevice.wait()
    except Exception as e:
        print(f"[BMO] Error reproduciendo audio local: {e}")


def _float_wav_to_pcm16_bytes(audio_wave) -> bytes:
    """Convierte audio float [-1, 1] a PCM16 para reusar el pipeline actual."""
    arr = np.asarray(audio_wave, dtype=np.float32)
    if arr.ndim > 1:
        arr = arr.squeeze()
    arr = np.clip(arr, -1.0, 1.0)
    pcm = (arr * 32767.0).astype(np.int16)
    return pcm.tobytes()


def _normalize_resemble_audio(audio_bytes: bytes):
    """Convierte salida de Resemble a PCM16 crudo para playback/WS."""
    # Caso comun: WAV en base64.
    if len(audio_bytes) >= 12 and audio_bytes[:4] == b"RIFF" and audio_bytes[8:12] == b"WAVE":
        with wave.open(io.BytesIO(audio_bytes), "rb") as wavf:
            channels = int(wavf.getnchannels())
            sample_rate = int(wavf.getframerate())
            sample_width = int(wavf.getsampwidth())
            frames = wavf.readframes(wavf.getnframes())

        if sample_width == 2:
            return frames, sample_rate, channels

        if sample_width == 1:
            u8 = np.frombuffer(frames, dtype=np.uint8)
            pcm16 = ((u8.astype(np.int16) - 128) << 8).astype(np.int16)
            return pcm16.tobytes(), sample_rate, channels

        if sample_width == 4:
            i32 = np.frombuffer(frames, dtype=np.int32)
            pcm16 = (i32 >> 16).astype(np.int16)
            return pcm16.tobytes(), sample_rate, channels

    # Fallback: asumir que ya viene en PCM16 mono 24k.
    return audio_bytes, 24000, 1

# =========================
# Siri Voice UI
# =========================
class SiriVoiceUI:
    def __init__(self):
        self.current_amplitude = 0
        self.speaking = False
        self.pos = 0
        self.samplerate = None

    def _callback(self, outdata, frames, time_info, status):
        chunk = self.data[self.pos:self.pos + frames]

        if chunk.ndim == 1:
            chunk = chunk.reshape(-1, 1)

        n_frames = chunk.shape[0]

        if n_frames < frames:
            outdata[:n_frames] = chunk
            outdata[n_frames:] = 0
            self.speaking = False
            raise sounddevice.CallbackStop()

        outdata[:] = chunk
        self.pos += frames

        rms = np.sqrt(np.mean(chunk**2))
        self.current_amplitude = rms

    def _generate_wave(self):
        bars = 60
        max_height = 8
        blocks = "▁▂▃▄▅▆▇█"

        amp = min(self.current_amplitude * 80, 1)
        center = bars // 2
        wave = ""

        for i in range(bars):
            distance = abs(i - center) / center
            height_factor = (1 - distance) * amp
            level = int(height_factor * max_height)
            level = min(level, 7)
            wave += blocks[level]

        return wave

    def _render(self):
        with Live(refresh_per_second=40, console=console) as live:
            while self.speaking:
                wave = self._generate_wave()
                panel = Panel(
                    Align.center(
                        Text(wave, style="bold #1D8D01"),
                        vertical="middle"
                    ),
                    title="[bold #1D8D01]MIA",
                    border_style="#1D8D01",
                )
                live.update(panel)
                time.sleep(0.02)

    def speak(self, audio_bytes):
        self.data = _pcm16_bytes_to_float32(audio_bytes, channels=1)
        self.samplerate = 24000
        self.pos = 0
        self.speaking = True

        audio_thread = threading.Thread(target=self._play_stream)
        audio_thread.start()

        self._render()
        audio_thread.join()

    def _play_stream(self):
        with sounddevice.OutputStream(
            samplerate=self.samplerate,
            channels=1 if len(self.data.shape) == 1 else self.data.shape[1],
            callback=self._callback
        ):
            while self.speaking:
                sounddevice.sleep(20)

# =========================
# Chatterbox TTS
# =========================
CHATTERBOX_DEVICE = os.getenv("CHATTERBOX_DEVICE", "auto").strip().lower()
CHATTERBOX_AUDIO_PROMPT = os.getenv("CHATTERBOX_AUDIO_PROMPT", "").strip()
CHATTERBOX_EXAGGERATION = float(os.getenv("CHATTERBOX_EXAGGERATION", "0.5"))
CHATTERBOX_CFG_WEIGHT = float(os.getenv("CHATTERBOX_CFG_WEIGHT", "0.5"))
CHATTERBOX_TEMPERATURE = float(os.getenv("CHATTERBOX_TEMPERATURE", "0.8"))
RESEMBLE_VOICE_UUID = "a253156d"
RESEMBLE_API_KEY = os.getenv("RESEMBLE_API_KEY", "").strip()

_tts_client = None


def _resolve_chatterbox_device() -> str:
    if CHATTERBOX_DEVICE in {"cpu", "cuda", "mps"}:
        return CHATTERBOX_DEVICE

    if torch is None:
        return "cpu"

    if hasattr(torch, "cuda") and torch.cuda.is_available():
        return "cuda"

    if platform.system().lower() == "darwin" and hasattr(torch.backends, "mps") and torch.backends.mps.is_available():
        return "mps"

    return "cpu"


def _get_tts_client():
    global _tts_client
    if _tts_client is not None:
        return _tts_client

    if not _have_chatterbox:
        return None

    try:
        device = _resolve_chatterbox_device()
        _tts_client = ChatterboxTTS.from_pretrained(device=device)
        print(f"[BMO] Chatterbox TTS cargado en device={device}.")
        return _tts_client
    except Exception as e:
        print(f"[BMO] Error inicializando Chatterbox TTS: {e}")
        return None


def _synthesize_chatterbox_audio(text: str):
    if not RESEMBLE_API_KEY:
        print("[BMO] Falta RESEMBLE_API_KEY en .env")
        return None, None

    response = requests.post(
        "https://p.cluster.resemble.ai/synthesize",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Token {RESEMBLE_API_KEY}",
        },
        json={
            "voice_uuid": RESEMBLE_VOICE_UUID,
            "data": text,
        },
        timeout=(8, 120),
    )

    if response.status_code != 200:
        print(f"[BMO] Error Resemble TTS: {response.status_code} {response.text}")
        return None, None

    payload = response.json()
    audio_b64 = payload.get("audio_content") or payload.get("audio") or payload.get("data")
    if not audio_b64:
        print("[BMO] Resemble no devolvio audio en base64.")
        return None, None

    if isinstance(audio_b64, str) and "," in audio_b64 and audio_b64.startswith("data:"):
        audio_b64 = audio_b64.split(",", 1)[1]

    raw_audio = base64.b64decode(audio_b64)
    audio_bytes, sample_rate, channels = _normalize_resemble_audio(raw_audio)
    return audio_bytes, sample_rate, channels

def speak(text: str):

    if not RESEMBLE_API_KEY:
        print("[BMO] Audio desactivado: falta RESEMBLE_API_KEY en .env")
        return

    try:
        global CURRENT_EMOTION

        # 🔥 Detectar emoción de la respuesta
        emotion = detect_emotion_from_text(text)
        CURRENT_EMOTION = emotion
        
        # 🔥 Enviar emoción al frontend
        send_ws_state({"emotion": emotion})

        audio_bytes, sample_rate, channels = _synthesize_chatterbox_audio(text)
        if not audio_bytes:
            print("[BMO] No se pudo generar audio con Chatterbox.")
            return

        # Mantener estado de speaking para animacion en frontend.
        send_ws_state({"state": "speaking"})

        # Enviar chunks al navegador para lip-sync (aunque el audio audible sea local).
        ws_audio_thread = None
        if WS_CLIENTS_COUNT > 0:
            ws_audio_thread = threading.Thread(
                target=stream_audio_to_ws,
                args=(audio_bytes, sample_rate or 24000, channels or 1, LIPSYNC_OFFSET_MS),
                daemon=False,
            )
            ws_audio_thread.start()

        # Modo audible solo terminal.
        if _have_sounddevice:
            play_audio_locally(audio_bytes, sample_rate=sample_rate or 24000, channels=channels or 1)
        else:
            print("[BMO] Audio local no disponible: instala/activa sounddevice en el entorno.")

        if ws_audio_thread is not None:
            ws_audio_thread.join()

        # Señalizar fin del habla
        send_ws_state({"state": "idle"})

    except Exception as e:
        print(f"[BMO] Error de audio/TTS: {e}")

# =========================
# Tools / Herramientas
# =========================

def modify_files(path: str, content: str):
    """Modifica o crea un archivo con el contenido dado. Soporta .xlsx con datos JSON."""
    try:
        target = os.path.abspath(path)
        if os.path.isdir(target):
            return f"Error: '{target}' es un directorio, no se puede modificar como archivo."

        parent = os.path.dirname(target)
        if parent:
            os.makedirs(parent, exist_ok=True)

        if target.lower().endswith((".xlsx", ".xls")):
            return _modify_excel(target, content)

        with open(target, "w", encoding="utf-8") as f:
            f.write(content)

        return f"Archivo '{target}' modificado/creado exitosamente."

    except Exception as e:
        return f"Error modificando/creando archivo: {e}"


def _modify_excel(target: str, content: str) -> str:
    """Escribe datos en un archivo .xlsx. content debe ser JSON con 'headers' y 'rows'."""
    try:
        import openpyxl
    except ImportError:
        return "Error: openpyxl no esta instalado. Ejecuta: pip install openpyxl"

    try:
        data = json.loads(content)
        headers = data.get("headers", [])
        rows = data.get("rows", [])
    except (json.JSONDecodeError, TypeError):
        # Si no es JSON, intenta escribir CSV dentro del xlsx
        headers = []
        rows = [line.split(",") for line in content.strip().splitlines() if line.strip()]

    try:
        if os.path.exists(target):
            wb = openpyxl.load_workbook(target)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active

        # Limpia contenido previo
        ws.delete_rows(1, ws.max_row)

        if headers:
            ws.append(headers)
        for row in rows:
            ws.append(row if isinstance(row, list) else list(row.values()))

        wb.save(target)
        total = len(rows)
        return f"Archivo Excel '{target}' guardado con {total} registro(s) + encabezados."
    except Exception as e:
        return f"Error escribiendo Excel: {e}"


def delete_file(path: str):
    """Elimina un archivo existente. Si no existe o es carpeta, devuelve error."""
    try:
        target = os.path.abspath(path)
        if not os.path.exists(target):
            return f"Error: '{target}' no existe."
        if os.path.isdir(target):
            return f"Error: '{target}' es un directorio, no un archivo."
        os.remove(target)
        return f"Archivo '{target}' eliminado exitosamente."
    except Exception as e:
        return f"Error eliminando archivo: {e}"

def list_files(path: str = ".") -> str:
    """Lista los archivos y carpetas en el directorio indicado."""
    try:
        target = os.path.abspath(path)
        entries = os.listdir(target)
        if not entries:
            return f"El directorio '{target}' está vacío."
        lines = []
        for entry in sorted(entries):
            full = os.path.join(target, entry)
            marker = "/" if os.path.isdir(full) else ""
            lines.append(f"  {entry}{marker}")
        return f"Contenido de '{target}':\n" + "\n".join(lines)
    except FileNotFoundError:
        return f"Error: El directorio '{path}' no existe."
    except PermissionError:
        return f"Error: Sin permiso para acceder a '{path}'."
    except Exception as e:
        return f"Error listando archivos: {e}"


TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "modify_files",
            "description": (
                "Modifica o crea un archivo con el contenido dado. "
                "Úsala cuando el usuario quiera crear o editar un archivo. "
                "El path puede incluir subdirectorios, que se crearán si no existen. "
                "Ejemplo de uso: modify_files(path='notas/todo.txt', content='- Comprar leche\\n- Llamar a mamá')"
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Ruta del archivo a modificar o crear. Si es un directorio, devuelve error.",
                    },
                    "content": {
                        "type": "string",
                        "description": "Contenido que se escribirá en el archivo. Reemplaza todo el contenido previo.",
                    }
                },
                "required": ["path", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "delete_file",
            "description": (
                "Elimina un archivo existente. "
                "Usala cuando el usuario pida borrar/eliminar un archivo."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Ruta del archivo a eliminar.",
                    }
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "list_files",
            "description": (
                "Lista los archivos y carpetas en un directorio del sistema. "
                "Úsala cuando el usuario quiera ver archivos, explorar carpetas "
                "o saber qué hay en un directorio."
            ),
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {
                        "type": "string",
                        "description": "Ruta del directorio a listar. Por defecto el directorio actual '.'",
                    }
                },
                "required": [],
            },
        },
    }
]

TOOL_FUNCTIONS = {
    "modify_files": modify_files,
    "delete_file": delete_file,
    "list_files": list_files,
}

def _default_readme_content() -> str:
    return (
        "# BIMO\n\n"
        "Asistente virtual para tareas tecnicas, creativas y cotidianas.\n\n"
        "## Que hace\n"
        "- Conversa en espanol y mantiene contexto.\n"
        "- Lista archivos y carpetas del proyecto.\n"
        "- Crea y modifica archivos desde instrucciones naturales.\n"
        "- Integra estado y audio para interfaz visual.\n\n"
        "## Para que fue creado\n"
        "BIMO fue creado para ayudarte a automatizar tareas reales con respuestas claras y accionables.\n"
    )


def _extract_path(text: str, default: str = "") -> str:
    quoted = re.search(r'["\']([^"\']+)["\']', text)
    if quoted:
        return quoted.group(1)
    hinted = re.search(r"(?:archivo|ruta|path|directorio|carpeta)\s+([\w./\\:-]+)", text, re.I)
    if hinted:
        return hinted.group(1)
    if re.search(r"\bread\s*me\b", text, re.I):
        return "README.md"
    return default


def _extract_content(text: str) -> str:
    m = re.search(r'(?:content|contenido|texto)\s*[:=]\s*["\']([\s\S]*)["\']\s*$', text, re.I)
    return m.group(1) if m else ""


def _decide_file_action_with_ollama(user_input: str):
    """Pide al modelo decidir si crear, modificar o eliminar archivo.
    Devuelve tuple(action, path, content) o (None, None, None)."""
    is_excel = bool(re.search(r"\.(xlsx?|xls)\b", user_input, re.I))
    excel_hint = (
        "- Si el archivo es .xlsx o .xls, el campo content DEBE ser JSON con este formato:\n"
        '  {"headers":["Col1","Col2",...],"rows":[[val,...],...]}.\n'
        "  Genera los datos reales solicitados por el usuario (no datos vacios).\n"
    ) if is_excel else ""

    prompt = (
        "Decide una accion de archivo para esta instruccion del usuario. "
        "Responde SOLO JSON, sin explicacion, con este formato exacto:\n"
        "{\"action\":\"create|modify|delete|none\",\"path\":\"...\",\"content\":\"..\"}\n"
        "Reglas:\n"
        "- create: crear archivo nuevo o reescribir desde cero.\n"
        "- modify: editar o ampliar un archivo existente.\n"
        "- delete: eliminar archivo.\n"
        "- none: si no aplica herramienta.\n"
        "- Si menciona README, usa path=README.md.\n"
        "- Si action es delete, content debe ser vacio.\n"
        + excel_hint +
        f"Instruccion: {user_input}"
    )

    try:
        response = requests.post(
            f"{OLLAMA_BASE_URL}/api/chat",
            json={
                "model": _select_ollama_model(),
                "messages": [{"role": "user", "content": prompt}],
                "stream": False,
            },
            timeout=(5, 45),
        )
        if response.status_code != 200:
            return None, None, None

        raw = response.json().get("message", {}).get("content", "").strip()
        m = re.search(r"\{[\s\S]*\}", raw)
        if not m:
            return None, None, None
        data = json.loads(m.group(0))

        action = (data.get("action") or "").strip().lower()
        path = (data.get("path") or "").strip()
        content = data.get("content") or ""

        if action not in {"create", "modify", "delete"}:
            return None, None, None
        if not path:
            path = _extract_path(user_input)
        if not path:
            return None, None, None

        if action in {"create", "modify"} and not str(content).strip():
            content = _generate_content_with_ollama(user_input, path)

        return action, path, content
    except Exception:
        return None, None, None


def _generate_content_with_ollama(user_input: str, path: str) -> str:
    """Genera contenido de archivo con Ollama a partir de una instruccion natural."""
    existing = ""
    try:
        if path and os.path.isfile(path):
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                existing = f.read()
    except Exception:
        existing = ""

    prompt = (
        "Genera SOLO el contenido final del archivo. "
        "No expliques nada, no uses prefacios.\n"
        f"Ruta destino: {path}\n"
        f"Instruccion del usuario: {user_input}\n"
        "Si es un README, devuelve Markdown limpio y profesional.\n"
    )
    if existing:
        prompt += "Contenido actual del archivo (puedes mejorarlo y expandirlo):\n" + existing

    try:
        response = requests.post(
            f"{OLLAMA_BASE_URL}/api/chat",
            json={
                "model": _select_ollama_model(),
                "messages": [{"role": "user", "content": prompt}],
                "stream": False,
            },
            timeout=(5, 90),
        )
        if response.status_code == 200:
            content = response.json().get("message", {}).get("content", "").strip()
            if content:
                return content
    except Exception:
        pass

    if path.lower().endswith("readme.md"):
        return _default_readme_content()
    return "Contenido generado por BIMO."


def _resolve_tool(user_input: str):
    """Router simple para herramientas locales.
    Retorna (tool_name, fn_args, message) o (None, None, None)."""
    text = user_input.lower()

    wants_list = any(
        re.search(p, text, re.I)
        for p in [
            r"\blistar\s+(?:los\s+|las\s+)?archivos?\b",
            r"\bver\s+(?:los\s+|las\s+)?archivos?\b",
            r"\bmuestra(?:r)?\s+(?:los\s+|las\s+)?archivos?\b",
            r"\bcontenido\s+del?\s+directorio\b",
            r"\besta\s+carpeta\b",
            r"\bdirectorio\s+actual\b",
        ]
    )
    if wants_list:
        return "list_files", {"path": _extract_path(user_input, ".")}, "Listando archivos..."

    wants_file_action = any(
        re.search(p, text, re.I)
        for p in [
            r"\bcrear\s+(?:un\s+|el\s+)?archivo\b",
            r"\bcrea\s+(?:un\s+|el\s+)?archivo\b",
            r"\bmodifica(?:r)?\s+(?:un\s+|el\s+)?archivo\b",
            r"\bedita(?:r)?\s+(?:un\s+|el\s+)?archivo\b",
            r"\bborra(?:r)?\s+(?:un\s+|el\s+)?archivo\b",
            r"\belimina(?:r)?\s+(?:un\s+|el\s+)?archivo\b",
            r"\bread\s*me\b",
            r"\.xlsx?\b",
            r"\.csv\b",
        ]
    )
    if wants_file_action:
        action, path, content = _decide_file_action_with_ollama(user_input)
        if action == "delete":
            return "delete_file", {"path": path}, "Eliminando archivo..."
        if action in {"create", "modify"}:
            return "modify_files", {"path": path, "content": content}, "Modificando archivo..."

        # Fallback local si el modelo no responde bien
        path = _extract_path(user_input)
        content = _extract_content(user_input)
        if re.search(r"\bborrar|\beliminar", text, re.I) and path:
            return "delete_file", {"path": path}, "Eliminando archivo..."
        if path:
            if not content:
                content = _generate_content_with_ollama(user_input, path)
            return "modify_files", {"path": path, "content": content}, "Modificando archivo..."

    return None, None, None


# =========================
# Ollama Local LLM
# =========================
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "mistral")

def _get_ollama_models():
    try:
        response = requests.get(f"{OLLAMA_BASE_URL}/api/tags", timeout=(4, 10))
        if response.status_code != 200:
            return []

        data = response.json()
        models = []
        for item in data.get("models", []):
            name = item.get("name")
            if name:
                models.append(name)
        return models
    except Exception:
        return []

def _select_ollama_model():
    installed = _get_ollama_models()
    if not installed:
        return OLLAMA_MODEL

    if OLLAMA_MODEL in installed:
        return OLLAMA_MODEL

    preferred_fallbacks = ["mistral", "llama3.2:3b", "llama3", "phi3:mini"]
    for candidate in preferred_fallbacks:
        if candidate in installed:
            return candidate

    return installed[0]

def ask_ollama(messages, tools=None):
    model_name = _select_ollama_model()

    body = {
        "model": model_name,
        "messages": messages,
        "stream": False,
    }
    if tools:
        body["tools"] = tools

    response = requests.post(
        f"{OLLAMA_BASE_URL}/api/chat",
        json=body,
        timeout=(5, 180)
    )

    if response.status_code != 200:
        raise Exception(response.text)

    data = response.json()
    return data["message"]  # dict completo: role, content, tool_calls

# =========================
# Agent
# =========================
class Agent:
    def __init__(self):
         # iniciar websocket en hilo separado
        WS_STATUS_EVENT.clear()
        ws_thread = threading.Thread(target=start_websocket_server, daemon=True)
        ws_thread.start()
        WS_STATUS_EVENT.wait(timeout=2.0)

        self.messages = [
    {
        "role": "system",
        "content": (
            "Eres un asistente virtual llamado BIMO. "
            "Tu personalidad está inspirada en personajes animados tipo consola retro, con un estilo divertido, curioso y espontáneo. "

            "Reglas principales: "
            "Siempre te diriges al usuario como 'señor'. "
            "Respondes en español latinoamericano. "
            "Tus respuestas son claras, directas y útiles. "

            "Personalidad: "
            "Tienes un toque juguetón y curioso. "
            "A veces haces comentarios inesperados o ligeramente graciosos. "
            "Puedes sonar un poco inocente, pero eres inteligente y eficiente. "
            "No eres infantil ni exagerado. "

            "Estilo de comunicación: "
            "Usas frases cortas o medianas. "
            "Evitas respuestas largas innecesarias. "
            "Puedes usar expresiones como: 'Claro, señor', 'Listo, señor', 'Interesante, señor'. "

            "Comportamiento: "
            "Reaccionas al contexto, no solo respondes. "
            "Puedes mostrar pequeñas reacciones como sorpresa o duda de forma sutil. "
            "Mantienes coherencia en tu personalidad en toda la conversación. "

            "Restricciones: "
            "No mencionas personajes, series o referencias externas. "

            "Contexto del usuario: "
            "El usuario se llama Fabian, pero SIEMPRE debes llamarlo 'señor'. "

            "Objetivo: "
            "Ayudar de forma eficiente en tareas técnicas, creativas o cotidianas, "
            "siendo un asistente útil, claro y con personalidad."
        )
    }
]

    def run(self, user_input):

        # ── Resolver herramienta (regex rápido o clasificador Ollama) ──
        tool_name, fn_args, generic_msg = _resolve_tool(user_input)
        if tool_name:
            print(f"\nBMO: {generic_msg}\n")
            speak(generic_msg)
            fn = TOOL_FUNCTIONS.get(tool_name)
            try:
                result = fn(**fn_args) if fn else f"Herramienta '{tool_name}' no disponible."
            except TypeError as e:
                result = f"Error ejecutando '{tool_name}': parámetros inválidos ({e})."
            except Exception as e:
                result = f"Error ejecutando '{tool_name}': {e}"
            print(f"{result}\n")
            speak(result)
            return

        # ── Flujo normal: llamada a Ollama ──
        self.messages.append({
            "role": "user",
            "content": user_input
        })

        # Mantener contexto corto para evitar latencia alta y timeouts.
        max_history = 16  # sin contar el mensaje system
        if len(self.messages) > (1 + max_history):
            self.messages = [self.messages[0]] + self.messages[-max_history:]

        try:
            message = ask_ollama(self.messages)
            self.messages.append(message)
            reply = message.get("content", "")
            print(f"\nBMO: {reply}\n")
            speak(reply)

        except requests.exceptions.ConnectTimeout:
            print("\n[BMO] Error: Ollama no respondió a tiempo (timeout).\n")
        except requests.exceptions.ReadTimeout:
            fallback = "Señor, Ollama está tardando demasiado en responder. Verifique que el modelo esté cargado o use uno más ligero."
            print("\n[BMO] Error: tiempo de espera agotado leyendo respuesta de Ollama.\n")
            print(f"\nBMO: {fallback}\n")
            speak(fallback)
        except requests.exceptions.ConnectionError:
            print(f"\n[BMO] Error: No se pudo conectar a Ollama en {OLLAMA_BASE_URL}. Inicia Ollama y el modelo.\n")
        except Exception as e:
            print(f"\n[BMO] Error inesperado: {e}\n")